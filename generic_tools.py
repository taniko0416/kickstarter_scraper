#!/usr/bin/env python3

import gspread
from oauth2client.service_account import ServiceAccountCredentials
import numpy as np
from datetime import datetime
import openpyxl
from time import sleep





def writeDoodProductsOnGS(all_data,now):

        good_products = []
        workbook = openpyxl.load_workbook('./products_DB.xlsx')
        sheet = workbook.create_sheet(now)
        sheet.active = workbook.get_sheet_by_name(now)

        print(sheet)
        # sheet = workbook[].active

        row,col = all_data.shape

        for i in range(col):
                price = int(all_data[5,i])
                if price > 1000000:
                        good_products.append(all_data[:,i])

        num_good_products = len(good_products)

        last_row = sheet.max_row

        # テスト用
        # print("最終行は%d行目です" %last_row)

        for i in range(num_good_products):

                h = i + 1 + last_row
                n1 = 'A' + str(h)
                n2 = 'B' + str(h)
                n3 = 'C' + str(h)
                n4 = 'D' + str(h)
                n5 = 'E' + str(h)
                n6 = 'F' + str(h)
                n7 = 'G' + str(h)
                n8 = 'H' + str(h)
                n9 = 'I' + str(h)

                sheet[n1] = good_products[i][0]
                sheet[n2] = int(good_products[i][2])
                sheet[n3] = good_products[i][3]
                sheet[n4] = int(good_products[i][5])
                sheet[n5] = int(good_products[i][6])
                sheet[n7] = good_products[i][4]
                sheet[n8] = good_products[i][7]
                sheet[n9] = now
        workbook.save('products_DB.xlsx')
        return 0

# #Excel内の重複の整形をするための関数
# def cleaningData(now):

#         workbook = openpyxl.load_workbook('/Users/niitsukouhei/Desktop/OurProducts/AnalysingKS/products_DB.xlsx')
#         sheet = workbook.get_sheet_by_name(now)
#         last_row = sheet.max_row

#         for i in range(last_row,1,-1):

#                 search_target_product = sheet.cell(i,1).value

#                 for j in range(i-1,1,-1):

#                         compare_product = sheet.cell(j,1).value

#                         if search_target_product == compare_product:

#                                 sheet.cell(j,1).value = ''


#         for i in range(last_row,1,-1):

#                 if sheet.cell(i,1).value == '':
#                         sheet.delete_rows(i)

#         workbook.save('products_DB.xlsx')

#         return 0



def analysing_data():

    search_target_name_arr = []
    search_target_time_arr = []
    search_target_backer_arr = []
    search_target_name_arr2 = []
    search_target_time_arr2 = []
    search_target_backer_arr2 = []

    workbook = openpyxl.load_workbook('./products_DB.xlsx')
    sheetnames = workbook.get_sheet_names()
    last_sheet = sheetnames[len(sheetnames)-1]
    second_last_sheet = sheetnames[len(sheetnames)-2]

    target_sheet = workbook[last_sheet]
    last_row = target_sheet.max_row

    for i in range(2,last_row+1,1):

        search_target_product = target_sheet.cell(i,1).value
        search_target_backer = target_sheet.cell(i,5).value
        search_target_time = target_sheet.cell(i,9).value
        search_target_time = search_target_time.replace('  ',' ')
        search_target_time = search_target_time.replace('-',':')
        search_target_time = search_target_time.replace('.','-')
        search_target_time = datetime.strptime(search_target_time, '%Y-%m-%d %H:%M:%S')
        search_target_name_arr.append(search_target_product)
        search_target_backer_arr.append(search_target_backer)
        search_target_time_arr.append(search_target_time.timestamp())

    np_arr = np.array([search_target_name_arr,search_target_time_arr,search_target_backer_arr])


    target_sheet2 = workbook[second_last_sheet]
    last_row2 = target_sheet2.max_row

    for i in range(2,last_row2+1,1):

        search_target_product2 = target_sheet2.cell(i,1).value
        search_target_backer2 = target_sheet2.cell(i,5).value
        search_target_time2 = target_sheet2.cell(i,9).value
        search_target_time2 = search_target_time2.replace('  ',' ')
        search_target_time2 = search_target_time2.replace('-',':')
        search_target_time2 = search_target_time2.replace('.','-')
        search_target_time2 = datetime.strptime(search_target_time2, '%Y-%m-%d %H:%M:%S')
        search_target_name_arr2.append(search_target_product2)
        search_target_backer_arr2.append(search_target_backer2)
        search_target_time_arr2.append(search_target_time2.timestamp())

    np_arr2 = np.array([search_target_name_arr2,search_target_time_arr2,search_target_backer_arr2])

    row,col = np_arr.shape
    row2,col2 = np_arr2.shape


    for i in range(col):
        product_name = str(np_arr[0,i])
        # print(str(np_arr[2,i]))
        for j in range(col2):
            product_name2 = str(np_arr2[0,j])

            if product_name == product_name2:
                # print(product_name)
                # print(product_name2)
                time = float(np_arr[1,i])
                time2 = float(np_arr2[1,j])
                d_time =  time - time2
                # print(d_time)
                backer = int(np_arr[2,i])
                backer2 = int(np_arr2[2,j])
                d_backer = backer - backer2
                num = i + 2
                # print(d_backer)
                n9 = 'J' + str(num)
                n10 = 'K' + str(num)
                n11 = 'E' + str(num)
                # print(num)
                target_sheet[n9] = d_time
                target_sheet[n10] = d_backer


        workbook.save('products_DB.xlsx')
